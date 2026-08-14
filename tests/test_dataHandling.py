
import csv
import json
import os

from dotenv import load_dotenv
from openpyxl import load_workbook
import pytest
from utils.jsonHandling import jsonHandling




# @pytest.mark.datahandling
def test_csvHandling():
    with open('testData/credentails.csv') as data:
            finalData = csv.DictReader(data)
            output = []
            for i in finalData:
                 output.append(i)

            print(output)
            # print(finalData)



def test_csvHandling_write():
    with open('testData/credentails.csv', mode='w',newline='') as data:
            finalData = csv.DictWriter(data, fieldnames=['username','password'])
            finalData.writeheader()
            finalData.writerow({'username':"tripur123_9",'password':"1234"})

def test_jsonHandling():
    file_path = 'testData/credentails.json'
    data = jsonHandling(file_path)

    assert 'positCreds' in data
    assert data['positCreds']['username'] == 'testuser'
    assert data['positCreds']['password'] == 'testpassword'
    assert data['negCreds']['username'] == 'testuser_neg'
    assert data['negCreds']['password'] == 'testpassword-Neg'

# pip install openpyxl
# @pytest.mark.datahandling
def test_excel():
     workbook =load_workbook('testData\\sample_creds.xlsx')
     sheet = workbook['sheet1']
     output = []
     for i in sheet.iter_rows(min_row=2, values_only=True):
          output.append(i)

     print(output)


# @pytest.mark.datahandling
def test_excel_write():
     workbook =load_workbook('testData\\sample_creds.xlsx')
     sheet = workbook['sheet1']
     # sheet.append(["tripur123_9","1234"])
     sheet["A7"] = "=SUM(A2:A6)"
     # sheet.delete_rows(2,sheet.max_row)
     workbook.save('testData\\sample_creds.xlsx')

# @pytest.mark.datahandling
def test_CLI():
     usname = os.getenv("usName_2july", "data1")
     pw = os.getenv("pw_2july","data2")
     print(usname)
     print(pw)

#pip install dotenv
@pytest.mark.datahandling
def test_env():
     load_dotenv(os.getenv('envfile'))
     usname = os.getenv("usName_3july")
     pw = os.getenv("pw_3july")
     url = os.getenv("url_3july")
     print(usname)
     print(pw)
     print(url)

          



    





